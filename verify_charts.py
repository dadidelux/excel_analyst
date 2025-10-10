"""Verify that charts were created in the Excel file"""
import openpyxl
import sys

file_path = sys.argv[1] if len(sys.argv) > 1 else 'outputs/analysis_report_20251009_125654.xlsx'

wb = openpyxl.load_workbook(file_path)
print(f'\n** Excel File Analysis: {file_path} **\n')
print(f'Sheets: {", ".join(wb.sheetnames)}')

# Check Charts sheet
if 'Charts' in wb.sheetnames:
    ws = wb['Charts']
    print(f'\nCharts Sheet Analysis:')
    print(f'  - Max row: {ws.max_row}')
    print(f'  - Max column: {ws.max_column}')

    # Check for native charts
    if hasattr(ws, '_charts'):
        print(f'  - Native charts found: {len(ws._charts)}')
        for idx, chart in enumerate(ws._charts):
            print(f'    Chart {idx + 1}: {type(chart).__name__} - "{chart.title}"')
    else:
        print('  - No native charts attribute found')

    # Check for images (embedded chart images)
    if hasattr(ws, '_images'):
        print(f'  - Embedded images found: {len(ws._images)}')
    else:
        print('  - No embedded images')

wb.close()
print('\nDone!')
