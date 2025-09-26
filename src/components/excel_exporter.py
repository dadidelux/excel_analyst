"""
Excel Export System

Creates professional Excel reports with multiple sheets, embedded charts,
formatted analysis results, and metadata summaries.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, LineChart, ScatterChart, PieChart, Reference
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime
import io
import base64
from PIL import Image as PILImage

from .data_analyzer import AnalysisResult
from .chart_generator import ChartOutput, ChartType
from .csv_loader import CSVMetadata
from ..utils.logger import get_logger
from ..utils.exceptions import ExcelExportError

logger = get_logger(__name__)


@dataclass
class ExcelExportConfig:
    """Configuration for Excel export"""
    include_raw_data: bool = True
    include_charts: bool = True
    include_summary: bool = True
    include_metadata: bool = True
    chart_size: Tuple[int, int] = (600, 400)
    max_rows_per_sheet: int = 1000000
    decimal_places: int = 2


@dataclass
class ExcelExportResult:
    """Result of Excel export operation"""
    file_path: str
    sheets_created: List[str]
    charts_embedded: int
    total_rows: int
    file_size_mb: float
    success: bool = True
    error_message: Optional[str] = None


class ExcelExporter:
    """
    Professional Excel export system with multi-sheet support and chart embedding
    """

    def __init__(self, output_dir: str = "outputs"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Excel styling definitions
        self.styles = self._define_styles()

    def _define_styles(self) -> Dict[str, Dict]:
        """Define consistent styling for Excel sheets"""
        return {
            'header': {
                'font': Font(bold=True, color='FFFFFF', size=12),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'subheader': {
                'font': Font(bold=True, color='000000', size=11),
                'fill': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'data': {
                'font': Font(color='000000', size=10),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'numeric': {
                'font': Font(color='000000', size=10),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'number_format': '#,##0.00'
            },
            'title': {
                'font': Font(bold=True, color='000000', size=14),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'insight': {
                'font': Font(color='000000', size=10, italic=True),
                'fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
                'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
            }
        }

    def export_analysis_to_excel(
        self,
        analysis_results: List[AnalysisResult],
        charts: List[ChartOutput],
        raw_data: Optional[pd.DataFrame] = None,
        metadata: Optional[CSVMetadata] = None,
        config: Optional[ExcelExportConfig] = None,
        filename: Optional[str] = None
    ) -> ExcelExportResult:
        """
        Export complete analysis to Excel with multiple sheets

        Args:
            analysis_results: List of analysis results
            charts: List of generated charts
            raw_data: Original dataset
            metadata: Dataset metadata
            config: Export configuration
            filename: Output filename

        Returns:
            ExcelExportResult with export details
        """
        if not config:
            config = ExcelExportConfig()

        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"analysis_report_{timestamp}.xlsx"

            file_path = self.output_dir / filename

            # Create workbook
            wb = Workbook()
            sheets_created = []
            charts_embedded = 0
            total_rows = 0

            # Remove default sheet
            wb.remove(wb.active)

            # Create Executive Summary sheet
            if config.include_summary:
                summary_sheet = self._create_summary_sheet(wb, analysis_results, metadata)
                sheets_created.append("Executive Summary")

            # Create Analysis Results sheets
            for i, result in enumerate(analysis_results):
                sheet_name = f"{result.analysis_type.replace(' ', '_')}"
                if len(sheet_name) > 31:  # Excel sheet name limit
                    sheet_name = sheet_name[:28] + f"_{i+1}"

                analysis_sheet = self._create_analysis_sheet(wb, result, sheet_name)
                sheets_created.append(sheet_name)
                total_rows += len(result.data) if isinstance(result.data, dict) else 0

            # Create Charts sheet
            if config.include_charts and charts:
                charts_sheet = self._create_charts_sheet(wb, charts)
                sheets_created.append("Charts")
                charts_embedded = len([c for c in charts if c.success])

            # Create Raw Data sheet
            if config.include_raw_data and raw_data is not None:
                raw_data_sheet = self._create_raw_data_sheet(wb, raw_data, config)
                sheets_created.append("Raw Data")
                total_rows += len(raw_data)

            # Create Metadata sheet
            if config.include_metadata and metadata:
                metadata_sheet = self._create_metadata_sheet(wb, metadata)
                sheets_created.append("Metadata")

            # Save workbook
            wb.save(file_path)

            # Calculate file size
            file_size_mb = file_path.stat().st_size / (1024 * 1024)

            logger.info(f"Excel export completed: {file_path}")

            return ExcelExportResult(
                file_path=str(file_path),
                sheets_created=sheets_created,
                charts_embedded=charts_embedded,
                total_rows=total_rows,
                file_size_mb=file_size_mb,
                success=True
            )

        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return ExcelExportResult(
                file_path="",
                sheets_created=[],
                charts_embedded=0,
                total_rows=0,
                file_size_mb=0,
                success=False,
                error_message=str(e)
            )

    def _create_summary_sheet(
        self,
        wb: Workbook,
        analysis_results: List[AnalysisResult],
        metadata: Optional[CSVMetadata]
    ) -> Any:
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary")

        # Title
        ws['A1'] = "Data Analysis Report - Executive Summary"
        ws.merge_cells('A1:F1')
        self._apply_style(ws['A1'], self.styles['title'])

        # Report metadata
        row = 3
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1

        if metadata:
            ws[f'A{row}'] = "Source File:"
            ws[f'B{row}'] = metadata.filename
            row += 1
            ws[f'A{row}'] = "Data Shape:"
            ws[f'B{row}'] = f"{metadata.shape[0]} rows × {metadata.shape[1]} columns"
            row += 2

        # Analysis summary
        ws[f'A{row}'] = "Analysis Summary"
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        # Headers
        headers = ['Analysis Type', 'Status', 'Key Findings', 'Insights Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['subheader'])
        row += 1

        # Analysis details
        for result in analysis_results:
            ws.cell(row=row, column=1, value=result.analysis_type)
            ws.cell(row=row, column=2, value="Success" if result.success else "Failed")
            ws.cell(row=row, column=3, value=result.summary[:100] + "..." if len(result.summary) > 100 else result.summary)
            ws.cell(row=row, column=4, value=len(result.insights))

            # Apply styling
            for col in range(1, 5):
                self._apply_style(ws.cell(row=row, column=col), self.styles['data'])
            row += 1

        # Key insights section
        row += 2
        ws[f'A{row}'] = "Key Insights"
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        # Collect all insights
        all_insights = []
        for result in analysis_results:
            for insight in result.insights:
                all_insights.append(f"• {insight}")

        for insight in all_insights[:10]:  # Top 10 insights
            ws[f'A{row}'] = insight
            self._apply_style(ws[f'A{row}'], self.styles['insight'])
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

        return ws

    def _create_analysis_sheet(
        self,
        wb: Workbook,
        analysis_result: AnalysisResult,
        sheet_name: str
    ) -> Any:
        """Create sheet for individual analysis result"""
        ws = wb.create_sheet(sheet_name)

        # Title
        ws['A1'] = f"{analysis_result.analysis_type} - Analysis Results"
        ws.merge_cells('A1:F1')
        self._apply_style(ws['A1'], self.styles['title'])

        row = 3

        # Summary
        ws[f'A{row}'] = "Summary:"
        ws[f'B{row}'] = analysis_result.summary
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        ws.merge_cells(f'B{row}:F{row}')
        row += 2

        # Analysis data
        if analysis_result.data:
            ws[f'A{row}'] = "Analysis Results"
            self._apply_style(ws[f'A{row}'], self.styles['header'])
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            # Convert analysis data to tabular format
            table_data = self._convert_analysis_data_to_table(analysis_result.data)

            if table_data:
                # Headers
                headers = list(table_data[0].keys()) if table_data else []
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    self._apply_style(cell, self.styles['subheader'])
                row += 1

                # Data rows
                for data_row in table_data:
                    for col, (key, value) in enumerate(data_row.items(), 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        if isinstance(value, (int, float)):
                            self._apply_style(cell, self.styles['numeric'])
                        else:
                            self._apply_style(cell, self.styles['data'])
                    row += 1

        # Insights section
        if analysis_result.insights:
            row += 1
            ws[f'A{row}'] = "Key Insights"
            self._apply_style(ws[f'A{row}'], self.styles['header'])
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            for insight in analysis_result.insights:
                ws[f'A{row}'] = f"• {insight}"
                self._apply_style(ws[f'A{row}'], self.styles['insight'])
                ws.merge_cells(f'A{row}:F{row}')
                row += 1

        # Metadata section
        if analysis_result.metadata:
            row += 1
            ws[f'A{row}'] = "Analysis Metadata"
            self._apply_style(ws[f'A{row}'], self.styles['header'])
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            for key, value in analysis_result.metadata.items():
                ws[f'A{row}'] = str(key).replace('_', ' ').title()
                ws[f'B{row}'] = str(value)
                self._apply_style(ws[f'A{row}'], self.styles['subheader'])
                self._apply_style(ws[f'B{row}'], self.styles['data'])
                row += 1

        self._auto_adjust_columns(ws)
        return ws

    def _create_charts_sheet(
        self,
        wb: Workbook,
        charts: List[ChartOutput]
    ) -> Any:
        """Create sheet with embedded charts"""
        ws = wb.create_sheet("Charts")

        # Title
        ws['A1'] = "Data Visualizations"
        ws.merge_cells('A1:F1')
        self._apply_style(ws['A1'], self.styles['title'])

        row = 3
        chart_count = 0

        for chart in charts:
            if not chart.success or not chart.base64_data:
                continue

            try:
                # Add chart title
                ws[f'A{row}'] = chart.title
                self._apply_style(ws[f'A{row}'], self.styles['subheader'])
                row += 1

                # Decode base64 image
                image_data = base64.b64decode(chart.base64_data)
                image_stream = io.BytesIO(image_data)

                # Create Excel image
                img = Image(image_stream)
                img.width = 600
                img.height = 400

                # Position image
                img.anchor = f'A{row}'
                ws.add_image(img)

                # Move to next position (leave space for image)
                row += 22  # Approximate rows for 400px height
                chart_count += 1

            except Exception as e:
                logger.warning(f"Failed to embed chart {chart.title}: {e}")
                ws[f'A{row}'] = f"Chart embedding failed: {chart.title}"
                row += 2

        if chart_count == 0:
            ws['A3'] = "No charts available for embedding"

        return ws

    def _create_raw_data_sheet(
        self,
        wb: Workbook,
        raw_data: pd.DataFrame,
        config: ExcelExportConfig
    ) -> Any:
        """Create sheet with raw data"""
        ws = wb.create_sheet("Raw Data")

        # Title
        ws['A1'] = "Original Dataset"
        ws.merge_cells('A1:F1')
        self._apply_style(ws['A1'], self.styles['title'])

        # Limit rows if necessary
        if len(raw_data) > config.max_rows_per_sheet:
            data_to_export = raw_data.head(config.max_rows_per_sheet)
            ws['A3'] = f"Note: Only first {config.max_rows_per_sheet} rows shown (total: {len(raw_data)})"
            start_row = 5
        else:
            data_to_export = raw_data
            start_row = 3

        # Export data
        for r_idx, row in enumerate(dataframe_to_rows(data_to_export, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Apply styling
                if r_idx == start_row:  # Header row
                    self._apply_style(cell, self.styles['subheader'])
                else:
                    if isinstance(value, (int, float)):
                        self._apply_style(cell, self.styles['numeric'])
                    else:
                        self._apply_style(cell, self.styles['data'])

        self._auto_adjust_columns(ws)
        return ws

    def _create_metadata_sheet(
        self,
        wb: Workbook,
        metadata: CSVMetadata
    ) -> Any:
        """Create sheet with dataset metadata"""
        ws = wb.create_sheet("Metadata")

        # Title
        ws['A1'] = "Dataset Metadata"
        ws.merge_cells('A1:F1')
        self._apply_style(ws['A1'], self.styles['title'])

        row = 3

        # Basic metadata
        metadata_items = [
            ("Filename", metadata.filename),
            ("Shape", f"{metadata.shape[0]} rows × {metadata.shape[1]} columns"),
            ("Encoding", metadata.encoding),
            ("Delimiter", metadata.delimiter),
            ("Memory Usage", f"{metadata.memory_usage:.2f} MB")
        ]

        for label, value in metadata_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            self._apply_style(ws[f'A{row}'], self.styles['subheader'])
            self._apply_style(ws[f'B{row}'], self.styles['data'])
            row += 1

        # Column information
        row += 2
        ws[f'A{row}'] = "Column Information"
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        # Column headers
        col_headers = ['Column Name', 'Data Type', 'Null Count', 'Null %']
        for col, header in enumerate(col_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_style(cell, self.styles['subheader'])
        row += 1

        # Column details
        total_rows = metadata.shape[0]
        for col_name in metadata.columns:
            dtype = metadata.dtypes.get(col_name, 'unknown')
            null_count = metadata.null_counts.get(col_name, 0)
            null_pct = (null_count / total_rows) * 100 if total_rows > 0 else 0

            ws.cell(row=row, column=1, value=col_name)
            ws.cell(row=row, column=2, value=str(dtype))
            ws.cell(row=row, column=3, value=null_count)
            ws.cell(row=row, column=4, value=f"{null_pct:.1f}%")

            for col in range(1, 5):
                if col in [3]:  # Numeric columns
                    self._apply_style(ws.cell(row=row, column=col), self.styles['numeric'])
                else:
                    self._apply_style(ws.cell(row=row, column=col), self.styles['data'])
            row += 1

        self._auto_adjust_columns(ws)
        return ws

    def _convert_analysis_data_to_table(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Convert analysis data dictionary to table format"""
        table_rows = []

        for key, value in data.items():
            if isinstance(value, dict):
                # Nested dictionary - flatten it
                for sub_key, sub_value in value.items():
                    table_rows.append({
                        'Category': key,
                        'Metric': sub_key,
                        'Value': sub_value
                    })
            elif isinstance(value, (list, tuple)):
                # List/tuple - enumerate items
                for i, item in enumerate(value):
                    table_rows.append({
                        'Category': key,
                        'Index': i,
                        'Value': item
                    })
            else:
                # Simple value
                table_rows.append({
                    'Category': key,
                    'Value': value
                })

        return table_rows

    def _apply_style(self, cell, style_dict):
        """Apply style dictionary to cell"""
        for attribute, value in style_dict.items():
            setattr(cell, attribute, value)

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width