"""Quick script to analyze the Excel output"""
import openpyxl
import sys

file_path = sys.argv[1] if len(sys.argv) > 1 else 'outputs/analysis_report_20251009_113947.xlsx'

wb = openpyxl.load_workbook(file_path)
print(f'Excel File Analysis: {file_path}')
print(f'\nSheets: {", ".join(wb.sheetnames)}')

print('\nSheet Details:')
for name in wb.sheetnames:
    ws = wb[name]
    print(f'  - {name}: {ws.max_row} rows x {ws.max_column} columns')

# Summary sheet
if 'Summary' in wb.sheetnames:
    print('\nSummary Sheet:')
    ws = wb['Summary']
    for i in range(1, min(25, ws.max_row + 1)):
        cell_a = ws.cell(i, 1).value
        cell_b = ws.cell(i, 2).value
        if cell_a:
            print(f'  {cell_a}: {cell_b}')

# Raw Data preview
if 'Raw Data' in wb.sheetnames:
    print('\nRaw Data Preview:')
    ws = wb['Raw Data']
    headers = [ws.cell(1, col).value for col in range(1, min(8, ws.max_column + 1))]
    print(f'  Columns: {", ".join(str(h) for h in headers if h)}')
    print(f'  Total rows: {ws.max_row - 1}')  # Excluding header

# Analysis sheets
analysis_sheets = [s for s in wb.sheetnames if 'Analysis' in s]
if analysis_sheets:
    print(f'\nAnalysis Sheets ({len(analysis_sheets)}):')
    for sheet in analysis_sheets:
        ws = wb[sheet]
        print(f'\n  {sheet}:')
        # Print first 10 rows
        for i in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(4, ws.max_column + 1)):
                val = ws.cell(i, col).value
                if val:
                    row_data.append(str(val)[:50])
            if row_data:
                print(f'    {" | ".join(row_data)}')

wb.close()
